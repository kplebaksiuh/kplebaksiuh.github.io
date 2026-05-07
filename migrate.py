"""
Skrip migrasi data dari Excel ke Supabase untuk Iuran Warga.
Jalankan: python migrate.py

Pastikan sudah install:
  pip install openpyxl supabase

Isi SUPABASE_URL dan SUPABASE_SERVICE_KEY di bawah.
Gunakan Service Role Key (bukan anon key) agar bisa bypass RLS.

Catatan:
- Sheet pengeluaran di Excel tidak ada kolom tanggal,
  sehingga tanggal akan diisi dengan nilai default.
  Setelah migrasi, edit tanggal via aplikasi jika diperlukan.
"""

import openpyxl
from supabase import create_client

# ===================================================
# KONFIGURASI - ISI SESUAI PROJECT SUPABASE ANDA
# ===================================================
SUPABASE_URL = "https://GANTI_DENGAN_URL_SUPABASE_ANDA.supabase.co"
SUPABASE_SERVICE_KEY = "GANTI_DENGAN_SERVICE_ROLE_KEY_ANDA"
EXCEL_PATH = r"c:\iuran_warga\data iuran warga.xlsx"
# ===================================================

sb = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)

# Urutan kolom: Okt(10), Nov(11), Des(12), Jan(1), Feb(2), Mar(3), Apr(4), Mei(5), Jun(6), Jul(7), Agu(8), Sep(9)
KOLOM_BULAN = [10, 11, 12, 1, 2, 3, 4, 5, 6, 7, 8, 9]


def get_tahun_id(label):
    res = sb.table("tahun_iuran").select("id").eq("label", label).single().execute()
    return res.data["id"]


def migrate_sheet(sheet, tahun_label):
    print(f"\n=== Migrasi sheet: {sheet.title} (Tahun {tahun_label}) ===")
    tahun_id = get_tahun_id(tahun_label)

    anggota_list = []
    pembayaran_list = []

    for row in sheet.iter_rows(min_row=4, values_only=True):
        no = row[0]
        nama = row[1]
        if not no or not nama or not isinstance(no, int):
            continue

        nama = str(nama).strip()
        if not nama:
            continue

        # Tentukan nominal_iuran dari nilai pertama yang ada di kolom bulan
        nominal_iuran = 5000
        for col_i in range(12):
            val = row[2 + col_i]
            if isinstance(val, (int, float)) and val > 0:
                nominal_iuran = int(val)
                break

        anggota_list.append({
            "nomor_urut": no,
            "nama": nama,
            "nominal_iuran": nominal_iuran,
            "is_active": True,
        })

    # Upsert anggota
    for anggota in anggota_list:
        sb.table("pengguna").upsert(anggota, on_conflict="nomor_urut").execute()

    # Ambil ulang pengguna (untuk dapat id)
    pengguna_db = sb.table("pengguna").select("id, nomor_urut").execute().data
    no_to_id = {p["nomor_urut"]: p["id"] for p in pengguna_db}

    print(f"  {len(anggota_list)} anggota diproses.")

    # Proses pembayaran
    for row in sheet.iter_rows(min_row=4, values_only=True):
        no = row[0]
        if not no or not isinstance(no, int):
            continue
        pengguna_id = no_to_id.get(no)
        if not pengguna_id:
            continue

        for col_i, bulan in enumerate(KOLOM_BULAN):
            val = row[2 + col_i]
            if isinstance(val, (int, float)) and val > 0:
                nominal = int(val)
                pembayaran_list.append({
                    "pengguna_id": pengguna_id,
                    "tahun_iuran_id": tahun_id,
                    "bulan": bulan,
                    "nominal": nominal,
                    "tanggal_bayar": f"{tahun_label}-01-01",  # tanggal estimasi
                })

    if pembayaran_list:
        sb.table("pembayaran").upsert(
            pembayaran_list,
            on_conflict="pengguna_id,tahun_iuran_id,bulan"
        ).execute()
        print(f"  {len(pembayaran_list)} record pembayaran dimigrasi.")


def migrate_pengeluaran(sheet):
    print("\n=== Migrasi Pengeluaran ===")
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        no = row[0]
        keterangan = row[1]
        nominal = row[2]
        if not isinstance(no, int) or not keterangan or not isinstance(nominal, (int, float)):
            continue
        rows.append({
            "tanggal": "2024-10-01",   # Tidak ada tanggal di Excel, gunakan default
            "keterangan": str(keterangan).strip(),
            "nominal": int(nominal),
        })
    if rows:
        sb.table("pengeluaran").insert(rows).execute()
        print(f"  {len(rows)} pengeluaran dimigrasi.")
    print("  Catatan: Tanggal pengeluaran menggunakan default (2024-10-01).")
    print("  Silakan edit tanggal masing-masing via aplikasi jika diperlukan.")


if __name__ == "__main__":
    import warnings
    warnings.filterwarnings("ignore")

    print("Membuka file Excel...")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    print(f"Sheet ditemukan: {wb.sheetnames}")

    migrate_sheet(wb["PEMASUKAN IURAN 2024"], "2024")
    migrate_sheet(wb["PEMASUKAN IURAN 2025"], "2025")
    migrate_pengeluaran(wb["DATA PENGELUARAN"])

    print("\n✓ Migrasi selesai!")
    print("Buka Supabase Dashboard untuk verifikasi data.")
